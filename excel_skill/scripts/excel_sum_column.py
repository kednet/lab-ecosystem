"""Add a column with a formula combining two existing columns.

Default operation: =COL1 * COL2 (e.g. qty * price = sum).
Use --op add for sums, --op formula to provide a custom formula template
with placeholders {A} and {B} (will be replaced with the actual cell refs).

Usage:
    python excel_sum_column.py FILE --sheet S --col1 G --col2 H --out-col I --header "Сумма"
    python excel_sum_column.py FILE --sheet S --col1 G --col2 H --op add --out-col I
    python excel_sum_column.py FILE --sheet S --col1 G --col2 H --op formula --formula "=ROUND({A}*{B},2)" --out-col I

By default creates 'FILE — out.xlsx' (use --in-place to overwrite).
"""
import argparse
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def col_to_idx(s: str) -> int:
    return int(s) if s.isdigit() else column_index_from_string(s)


def main() -> int:
    ap = argparse.ArgumentParser(description="Add a formula column")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--col1", required=True, help="First column (e.g. G or 7)")
    ap.add_argument("--col2", required=True, help="Second column")
    ap.add_argument("--out-col", required=True, help="Where to write the formula (e.g. I or 9)")
    ap.add_argument("--header", default="", help="Header for the new column")
    ap.add_argument("--op", choices=("mul", "add", "sub", "div", "formula"),
                    default="mul", help="Default: mul. 'formula' uses --formula template")
    ap.add_argument("--formula", help='Formula template with {A} and {B}, e.g. "=ROUND({A}*{B},2)"')
    ap.add_argument("--format", default="#,##0.00", help="Number format for the new column")
    ap.add_argument("--width", type=int, default=14, help="Column width")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--out", type=Path)
    args = ap.parse_args()

    c1 = col_to_idx(args.col1)
    c2 = col_to_idx(args.col2)
    out_idx = col_to_idx(args.out_col)
    out_letter = get_column_letter(out_idx)
    l1 = get_column_letter(c1)
    l2 = get_column_letter(c2)

    if args.op == "formula" and not args.formula:
        ap.error("--formula required for --op formula")

    if args.in_place:
        target = args.file
    else:
        target = args.out or args.file.with_name(args.file.stem + " — out.xlsx")
        if target == args.file:
            target = args.file.with_name(args.file.stem + " — out.xlsx")
        shutil.copy2(args.file, target)

    try:
        wb = openpyxl.load_workbook(target)
    except PermissionError:
        print(f"❌ {target.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    if args.header:
        ws.cell(row=1, column=out_idx).value = args.header

    op_templates = {
        "mul": "={A}*{B}",
        "add": "={A}+{B}",
        "sub": "={A}-{B}",
        "div": "=IFERROR({A}/{B},0)",
    }
    template = args.formula if args.op == "formula" else op_templates[args.op]

    filled = 0
    for r in range(2, ws.max_row + 1):
        v1 = ws.cell(row=r, column=c1).value
        v2 = ws.cell(row=r, column=c2).value
        if v1 is None or v2 is None:
            continue
        formula = template.format(A=f"{l1}{r}", B=f"{l2}{r}")
        cell = ws[f"{out_letter}{r}"]
        cell.value = formula
        cell.number_format = args.format
        filled += 1

    ws.column_dimensions[out_letter].width = args.width
    wb.save(target)
    print(f"✅ Записано {filled} формул в колонку {out_letter} (шапка: '{args.header or '—'}')")
    print(f"→ сохранено: {target}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
