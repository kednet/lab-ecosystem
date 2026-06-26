"""Count unique values in a column.

Usage:
    python excel_count_unique.py FILE --sheet S --col F
    python excel_count_unique.py FILE --sheet S --col F --top 10        # show top 10 most common
    python excel_count_unique.py FILE --sheet S --col F --case-insensitive
"""
import argparse
import sys
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def col_to_idx(s: str) -> int:
    return int(s) if s.isdigit() else column_index_from_string(s)


def main() -> int:
    ap = argparse.ArgumentParser(description="Count unique values in a column")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--col", required=True)
    ap.add_argument("--top", type=int, default=0, help="Show top N most common values")
    ap.add_argument("--case-insensitive", action="store_true")
    args = ap.parse_args()

    col = col_to_idx(args.col)
    letter = get_column_letter(col)

    try:
        wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    except PermissionError:
        print(f"❌ {args.file.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2
    except FileNotFoundError:
        print(f"❌ Файл не найден: {args.file}", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    values: list = []
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if col - 1 >= len(row):
            skipped += 1
            continue
        v = row[col - 1]
        if v is None or (isinstance(v, str) and not v.strip()):
            skipped += 1
            continue
        if args.case_insensitive and isinstance(v, str):
            v = v.strip().lower()
        elif isinstance(v, str):
            v = v.strip()
        values.append(v)
    wb.close()

    counter = Counter(values)
    print(f"📊 Колонка {letter} листа '{args.sheet}':")
    print(f"   Всего непустых значений: {len(values)}")
    print(f"   Пропущено пустых: {skipped}")
    print(f"   Уникальных: {len(counter)}")
    if args.top and args.top > 0:
        print(f"\n   Топ-{min(args.top, len(counter))} по частоте:")
        for val, cnt in counter.most_common(args.top):
            v_str = str(val)[:60] + ('…' if len(str(val)) > 60 else '')
            print(f"     {cnt:>5}  {v_str}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
