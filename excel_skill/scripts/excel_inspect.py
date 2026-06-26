"""Inspect an .xlsx file: list sheets, dimensions, headers, first N rows.

Usage:
    python excel_inspect.py FILE [--sheet NAME] [--head N] [--out FILE.txt]

Examples:
    python excel_inspect.py "прайс.xlsx"
    python excel_inspect.py "отчёт.xlsx" --sheet "Январь" --head 20
"""
import argparse
import sys
from pathlib import Path

import openpyxl


def inspect(path: Path, sheet: str | None, head: int) -> int:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        print(f"❌ Не могу открыть {path.name} — файл залочен. Закрой его в Excel и попробуй снова.", file=sys.stderr)
        return 2
    except FileNotFoundError:
        print(f"❌ Файл не найден: {path}", file=sys.stderr)
        return 2

    print(f"📄 {path}")
    print(f"📑 Листы: {wb.sheetnames}")

    sheets = [sheet] if sheet else wb.sheetnames
    for s in sheets:
        if s not in wb.sheetnames:
            print(f"⚠️  Лист '{s}' не найден", file=sys.stderr)
            continue
        ws = wb[s]
        print(f"\n=== {s} ===")
        # Note: with read_only=True dimensions are accurate
        print(f"  Размер: {ws.max_row} rows × {ws.max_column} cols")
        # Headers
        try:
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            print("  (пустой лист)")
            continue
        print(f"  Заголовки (row 1): {header_row}")
        # First N data rows
        print(f"  Первые {head} строк данных:")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=1 + head, values_only=True), start=2):
            # truncate long cell values
            row_short = tuple(str(v)[:50] + ('…' if v is not None and len(str(v)) > 50 else '') for v in row)
            print(f"    r{i:>3}: {row_short}")
    wb.close()
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(description="Inspect an Excel file")
    ap.add_argument("file", type=Path, help="Path to .xlsx file")
    ap.add_argument("--sheet", help="Specific sheet name (default: all)")
    ap.add_argument("--head", type=int, default=5, help="Number of data rows to show (default: 5)")
    ap.add_argument("--out", type=Path, help="Save output to this file")
    args = ap.parse_args()

    import io
    buf = io.StringIO()
    old, sys.stdout = sys.stdout, buf
    try:
        rc = inspect(args.file, args.sheet, args.head)
    finally:
        sys.stdout = old
    text = buf.getvalue()
    print(text, end='')
    if args.out:
        args.out.write_text(text, encoding='utf-8')
        print(f"\n→ сохранено в {args.out}")
    return rc


if __name__ == "__main__":
    sys.exit(main())
