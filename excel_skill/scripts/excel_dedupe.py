"""Find (default) or delete duplicate rows by a key column.

Usage:
    python excel_dedupe.py FILE --sheet S --key-col F                 # show dupes
    python excel_dedupe.py FILE --sheet S --key-col F --delete       # remove dupes, keep first
    python excel_dedupe.py FILE --sheet S --key-col F --delete --keep last

By default creates 'FILE — out.xlsx'.
"""
import argparse
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def col_to_idx(s: str) -> int:
    return int(s) if s.isdigit() else column_index_from_string(s)


def main() -> int:
    ap = argparse.ArgumentParser(description="Find/delete duplicate rows by key column")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--key-col", required=True)
    ap.add_argument("--delete", action="store_true", help="Delete duplicate rows")
    ap.add_argument("--keep", choices=("first", "last"), default="first")
    ap.add_argument("--case-insensitive", action="store_true")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--out", type=Path)
    args = ap.parse_args()

    col = col_to_idx(args.key_col)
    letter = get_column_letter(col)

    if args.in_place:
        target = args.file
    else:
        target = args.out or args.file.with_name(args.file.stem + " — out.xlsx")
        if target == args.file:
            target = args.file.with_name(args.file.stem + " — out.xlsx")
        shutil.copy2(args.file, target)

    try:
        wb = openpyxl.load_workbook(target)
    except PermissionError:
        print(f"❌ {target.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    seen: dict = {}
    row_keys: list = []  # (row_num, key, key_normalized)
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col).value
        if v is None or (isinstance(v, str) and not v.strip()):
            continue
        norm = str(v).strip().lower() if args.case_insensitive else (str(v).strip() if isinstance(v, str) else v)
        row_keys.append((r, v, norm))

    groups: dict = defaultdict(list)
    for r, v, norm in row_keys:
        groups[norm].append((r, v))

    dupes = {k: rows for k, rows in groups.items() if len(rows) > 1}
    total_extra = sum(len(rows) - 1 for rows in dupes.values())

    if not dupes:
        print(f"✅ Дубликатов по колонке {letter} не найдено")
        return 0

    print(f"🔍 Найдено {len(dupes)} дубль-ключей ({total_extra} лишних строк):")
    for k, rows in list(dupes.items())[:20]:
        print(f"   • '{k}' — {len(rows)} строк: {[(r, str(v)[:30]) for r, v in rows]}")
    if len(dupes) > 20:
        print(f"   … и ещё {len(dupes) - 20}")

    if args.delete:
        rows_to_delete: set = set()
        for k, rows in dupes.items():
            rows_sorted = sorted(rows, key=lambda x: x[0])  # by row
            if args.keep == "last":
                rows_sorted = rows_sorted[::-1]
            for r, _ in rows_sorted[1:]:  # skip the kept one
                rows_to_delete.add(r)
        # delete from bottom to top to keep row numbers stable
        for r in sorted(rows_to_delete, reverse=True):
            ws.delete_rows(r, 1)
        wb.save(target)
        print(f"✅ Удалено {len(rows_to_delete)} строк, оставлено {len(dupes)} уникальных")
        print(f"→ сохранено: {target}")
    else:
        print("ℹ️  Используй --delete чтобы удалить дубликаты")

    return 0


if __name__ == "__main__":
    sys.exit(main())
