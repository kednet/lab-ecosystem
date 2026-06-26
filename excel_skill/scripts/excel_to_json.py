"""Export a sheet to JSON list of dicts.

First row is treated as header. Empty trailing rows are skipped.

Usage:
    python excel_to_json.py FILE --sheet S --out OUT.json
    python excel_to_json.py FILE --sheet S --out OUT.json --pretty
"""
import argparse
import json
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser(description="Export Excel sheet to JSON")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--pretty", action="store_true", help="Pretty-print (indent=2)")
    ap.add_argument("--key-as-text", action="store_true", help="Convert all values to strings")
    args = ap.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    except PermissionError:
        print(f"❌ {args.file.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2
    except FileNotFoundError:
        print(f"❌ Файл не найден: {args.file}", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        print("⚠️  Пустой лист", file=sys.stderr)
        args.out.write_text("[]", encoding="utf-8")
        return 0

    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        # skip rows that are entirely None
        if all(v is None for v in row):
            continue
        record = {}
        for h, v in zip(headers, row):
            if args.key_as_text and v is not None:
                v = str(v)
            record[h] = v
        result.append(record)

    indent = 2 if args.pretty else None
    args.out.write_text(json.dumps(result, indent=indent, ensure_ascii=False, default=str),
                        encoding="utf-8")
    print(f"✅ Записано {len(result)} записей в {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
