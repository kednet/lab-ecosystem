"""Export a sheet to CSV. UTF-8 BOM by default (so Excel opens it without garbled Cyrillic).

Usage:
    python excel_to_csv.py FILE --sheet S --out OUT.csv
    python excel_to_csv.py FILE --sheet S --out OUT.csv --encoding utf-8    # plain UTF-8
    python excel_to_csv.py FILE --sheet S --out OUT.csv --sep ";"          # for EU Excel

Default: utf-8-sig (BOM), separator ",".
"""
import argparse
import csv
import sys
from pathlib import Path

import openpyxl


def main() -> int:
    ap = argparse.ArgumentParser(description="Export Excel sheet to CSV")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--out", type=Path, required=True)
    ap.add_argument("--encoding", default="utf-8-sig",
                    help="File encoding (default: utf-8-sig for Excel compatibility)")
    ap.add_argument("--sep", default=",", help="Field separator (default: ',')")
    ap.add_argument("--no-header", action="store_true", help="Skip first row")
    args = ap.parse_args()

    try:
        wb = openpyxl.load_workbook(args.file, read_only=True, data_only=True)
    except PermissionError:
        print(f"❌ {args.file.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2
    except FileNotFoundError:
        print(f"❌ Файл не найден: {args.file}", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not args.no_header and rows:
        rows = rows[1:]

    with args.out.open("w", encoding=args.encoding, newline="") as f:
        writer = csv.writer(f, delimiter=args.sep)
        for row in rows:
            writer.writerow(["" if v is None else v for v in row])

    print(f"✅ Записано {len(rows)} строк в {args.out}")
    print(f"   encoding={args.encoding}, sep='{args.sep}'")
    return 0


if __name__ == "__main__":
    sys.exit(main())
