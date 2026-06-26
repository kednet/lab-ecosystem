"""Fill empty cells in a column with a value or with neighbour cell value.

Modes:
    previous   — copy the value from the previous non-empty cell above
    next       — copy the value from the next non-empty cell below
    value:X    — fill with literal value X (number, string, date in ISO format)
    default    — alias for 'previous'

Usage:
    python excel_fill_missing.py FILE --sheet S --col H --mode previous
    python excel_fill_missing.py FILE --sheet S --col H --mode value:0
    python excel_fill_missing.py FILE --sheet S --col H --mode value:"нет цены"

By default creates 'FILE — out.xlsx'.
"""
import argparse
import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def col_to_idx(s: str) -> int:
    return int(s) if s.isdigit() else column_index_from_string(s)


def coerce(value: str):
    # Try int, then float, then leave as string
    try:
        return int(value)
    except ValueError:
        pass
    try:
        return float(value)
    except ValueError:
        pass
    return value


def main() -> int:
    ap = argparse.ArgumentParser(description="Fill empty cells in a column")
    ap.add_argument("file", type=Path)
    ap.add_argument("--sheet", required=True)
    ap.add_argument("--col", required=True)
    ap.add_argument("--mode", required=True,
                    help="previous | next | value:X (e.g. value:0, value:нет цены)")
    ap.add_argument("--max-rows", type=int, help="Limit number of rows to process")
    ap.add_argument("--in-place", action="store_true")
    ap.add_argument("--out", type=Path)
    args = ap.parse_args()

    col = col_to_idx(args.col)
    letter = get_column_letter(col)

    if args.in_place:
        target = args.file
    else:
        target = args.out or args.file.with_name(args.file.stem + " — out.xlsx")
        if target == args.file:
            target = args.file.with_name(args.file.stem + " — out.xlsx")
        shutil.copy2(args.file, target)

    try:
        wb = openpyxl.load_workbook(target)
    except PermissionError:
        print(f"❌ {target.name} залочен. Закрой в Excel.", file=sys.stderr)
        return 2

    if args.sheet not in wb.sheetnames:
        print(f"❌ Лист '{args.sheet}' не найден", file=sys.stderr)
        return 2
    ws = wb[args.sheet]

    last_row = min(ws.max_row, 1 + (args.max_rows or ws.max_row))
    filled = 0

    if args.mode in ("previous", "default"):
        last_val = None
        for r in range(2, last_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None or (isinstance(v, str) and not v.strip()):
                if last_val is not None:
                    ws.cell(row=r, column=col).value = last_val
                    filled += 1
            else:
                last_val = v
    elif args.mode == "next":
        # walk backwards
        nxt_val = None
        for r in range(last_row, 1, -1):
            v = ws.cell(row=r, column=col).value
            if v is None or (isinstance(v, str) and not v.strip()):
                if nxt_val is not None:
                    ws.cell(row=r, column=col).value = nxt_val
                    filled += 1
            else:
                nxt_val = v
    elif args.mode.startswith("value:"):
        raw = args.mode[len("value:"):]
        fill_val = coerce(raw)
        for r in range(2, last_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None or (isinstance(v, str) and not v.strip()):
                ws.cell(row=r, column=col).value = fill_val
                filled += 1
    else:
        ap.error(f"Unknown mode: {args.mode}")

    wb.save(target)
    print(f"✅ Заполнено {filled} пустых ячеек в колонке {letter} (mode: {args.mode})")
    print(f"→ сохранено: {target}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
