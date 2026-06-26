"""VLOOKUP-style: pull values from one sheet/file into another by key.

Both source and destination are optional — supports single-file (two sheets)
or two-file mode. By default creates 'FILE — out.xlsx' alongside the source
(use --in-place to overwrite).

Usage:
    python excel_vlookup.py DST_FILE
        --dst-sheet SHEET --dst-key-col COL --dst-val-col COL
        --src-file SRC_FILE --src-sheet SHEET --src-key-col COL --src-val-col COL
        [--key-as-text] [--on-missing skip|raise]
        [--in-place | --out OUT_FILE]

If --src-file is omitted, source is the same file as destination.
"""
import argparse
import sys
from pathlib import Path

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def col_letter_to_index(s: str) -> int:
    s = s.strip()
    # accept "F" or "6"
    return int(s) if s.isdigit() else column_index_from_string(s)


def build_map(src_path: Path, src_sheet: str, src_key_col: int, src_val_col: int,
              key_as_text: bool) -> tuple[dict, int, int]:
    wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
    ws = wb[src_sheet]
    keymap: dict = {}
    dup = 0
    skipped = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if src_key_col - 1 >= len(row):
            continue
        key = row[src_key_col - 1]
        if key is None or (isinstance(key, str) and not key.strip()):
            skipped += 1
            continue
        if key_as_text:
            key = str(key).strip()
        if src_val_col - 1 >= len(row):
            val = None
        else:
            val = row[src_val_col - 1]
        if key in keymap:
            dup += 1
            continue
        keymap[key] = val
    wb.close()
    return keymap, dup, skipped


def fill_dst(dst_path: Path, dst_sheet: str, dst_key_col: int, dst_val_col: int,
             keymap: dict, key_as_text: bool, on_missing: str) -> tuple[int, int, list]:
    wb = openpyxl.load_workbook(dst_path)
    ws = wb[dst_sheet]
    filled, skipped_existing, missing = 0, 0, []
    key_letter = get_column_letter(dst_key_col)
    val_letter = get_column_letter(dst_val_col)
    for r in range(2, ws.max_row + 1):
        k = ws[f"{key_letter}{r}"].value
        if k is None or (isinstance(k, str) and not k.strip()):
            continue
        if key_as_text:
            k = str(k).strip()
        cur = ws[f"{val_letter}{r}"].value
        if cur is not None and cur != "":
            skipped_existing += 1
            continue
        if k in keymap and keymap[k] is not None:
            ws[f"{val_letter}{r}"] = keymap[k]
            filled += 1
        else:
            if on_missing == "raise":
                raise KeyError(f"Артикул '{k}' (row {r}) не найден в источнике")
            missing.append((r, k))
    return filled, skipped_existing, missing


def main() -> int:
    ap = argparse.ArgumentParser(description="VLOOKUP from one sheet/file to another")
    ap.add_argument("dst_file", type=Path, help="Destination .xlsx file (will be modified)")
    ap.add_argument("--dst-sheet", required=True)
    ap.add_argument("--dst-key-col", required=True, help="Letter or number, e.g. F or 6")
    ap.add_argument("--dst-val-col", required=True, help="Where to put the looked-up value")
    ap.add_argument("--src-file", type=Path, help="Source file (default: same as dst)")
    ap.add_argument("--src-sheet", help="Source sheet (default: same as dst)")
    ap.add_argument("--src-key-col", help="Key column in source (default: same as dst)")
    ap.add_argument("--src-val-col", help="Value column in source (default: same as dst)")
    ap.add_argument("--key-as-text", action="store_true", help="Force string comparison + strip")
    ap.add_argument("--on-missing", choices=("skip", "raise"), default="skip")
    ap.add_argument("--in-place", action="store_true", help="Overwrite dst file")
    ap.add_argument("--out", type=Path, help="Output path (default: 'dst — out.xlsx' next to dst)")
    args = ap.parse_args()

    src_path = args.src_file or args.dst_file
    src_sheet = args.src_sheet or args.dst_sheet
    dst_key = col_letter_to_index(args.dst_key_col)
    dst_val = col_letter_to_index(args.dst_val_col)
    src_key = col_letter_to_index(args.src_key_col or args.dst_key_col)
    src_val = col_letter_to_index(args.src_val_col or args.dst_val_col)

    if src_sheet not in openpyxl.load_workbook(src_path, read_only=True).sheetnames:
        print(f"❌ Лист '{src_sheet}' не найден в {src_path}", file=sys.stderr)
        return 2
    if args.dst_sheet not in openpyxl.load_workbook(args.dst_file, read_only=True).sheetnames:
        print(f"❌ Лист '{args.dst_sheet}' не найден в {args.dst_file}", file=sys.stderr)
        return 2

    keymap, dup, skipped_src = build_map(src_path, src_sheet, src_key, src_val, args.key_as_text)
    print(f"📥 Источник: {len(keymap)} уникальных ключей, {dup} дубликатов, {skipped_src} пустых строк пропущено")

    if args.in_place:
        out_path = args.dst_file
    else:
        out_path = args.out or args.dst_file.with_name(args.dst_file.stem + " — out.xlsx")
        if out_path == args.dst_file:
            print("⚠️  --in-place не задан, но имя совпадает с исходником. Использую '... — out.xlsx'.", file=sys.stderr)
            out_path = args.dst_file.with_name(args.dst_file.stem + " — out.xlsx")

    if not args.in_place:
        import shutil
        shutil.copy2(args.dst_file, out_path)

    try:
        filled, kept, missing = fill_dst(
            out_path, args.dst_sheet, dst_key, dst_val, keymap, args.key_as_text, args.on_missing
        )
    except PermissionError:
        print(f"❌ Файл {out_path.name} залочен. Закрой в Excel и попробуй снова.", file=sys.stderr)
        return 2

    print(f"✅ Заполнено: {filled} | Уже было: {kept} | Не найдено: {len(missing)}")
    if missing[:10]:
        for r, k in missing[:10]:
            print(f"   • row {r}: '{k}'")
        if len(missing) > 10:
            print(f"   … и ещё {len(missing) - 10}")
    print(f"→ сохранено: {out_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
